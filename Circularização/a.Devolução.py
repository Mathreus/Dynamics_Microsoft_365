import pyodbc
import pandas as pd
from datetime import datetime
import os

# Configurações de conexão com o banco de dados
def conectar_banco():
    """
    Estabelece conexão com o banco de dados SQL Server
    Ajuste os parâmetros conforme sua configuração
    """
    try:
        # Configurações de conexão 
        server = 'DCMDWF01A.MOURA.INT'  
        database = 'ax'   
        username = 'uAuditoria' 
        password = '@ud!t0$!@202&22'  
        
        # String de conexão
        conn_str = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'
        
        # Estabelecer conexão
        conexao = pyodbc.connect(conn_str)
        print("Conexão estabelecida com sucesso!")
        return conexao
        
    except Exception as e:
        print(f"Erro ao conectar ao banco de dados: {e}")
        return None

def executar_query(conexao, query):
    """
    Executa uma query SQL e retorna um DataFrame pandas
    """
    try:
        df = pd.read_sql(query, conexao)
        return df
    except Exception as e:
        print(f"Erro ao executar query: {e}")
        return pd.DataFrame()

def main():
    # Definir as queries
    query_devolucao = """
    SELECT
        COD_ESTABELECIMENTO,
        COD_CLIENTE,
        NOME_CLIENTE,
        SUM(QUANTIDADE) AS VOLUME_DEVOLVIDO,
        SUM(VALOR) AS VALOR_DEVOLVIDO
    FROM 
        VW_AUDIT_RM_ORDENS_VENDA
    WHERE
        COD_ESTABELECIMENTO = 'R121'
        AND DATA_NOTA_FISCAL BETWEEN '2025-07-07' AND '2026-01-07' 
        AND PARA_FATURAMENTO = 'SIM'
        AND CFOP IN ('1.201', '1.202', '1.203', '1.204', '1.410', '1.411', '1.553', '1.660', '1.661', '1.662', 
                    '2.201', '2.202', '2.203', '2.204', '2.410', '2.411', '2.553', '2.660', '2.661', '2.662',
                    '3.201', '3.202', '3.211', '3.553')
    GROUP BY
        COD_ESTABELECIMENTO,
        COD_CLIENTE,
        NOME_CLIENTE
    """

    query_faturamento = """
    SELECT
        COD_ESTABELECIMENTO,
        COD_CLIENTE,
        NOME_CLIENTE,
        SUM(QUANTIDADE) AS VOLUME_VENDAS,
        SUM(VALOR) AS VALOR_VENDAS
    FROM 
        VW_AUDIT_RM_ORDENS_VENDA
    WHERE 
        COD_ESTABELECIMENTO = 'R121' 
        AND DATA_NOTA_FISCAL BETWEEN '2025-07-07' AND '2026-01-07'  
        AND PARA_FATURAMENTO = 'Sim'
        AND CFOP IN ('5.100', '5.101', '5.102', '5.103', '5.104', '5.105', '5.106', '5.109', '5.110', '5.111', 
                    '5.112', '5.113', '5.114', '5.115', '5.116', '5.117', '5.118', '5.119', '5.120', '5.122', 
                    '5.123', '5.250','5.251', '5.252', '5.253', '5.254', '5.255', '5.256', '5.257', '5.258', 
                    '5.401', '5.402', '5.403', '5.405', '5.651', '5.652', '5.653', '5.654', '5.655', '5.656',
                    '5.667', '6.101', '6.102', '6.103','6.104', '6.105', '6.106', '6.107', '6.108', '6.109',
                    '6.110', '6.111', '6.112', '6.113', '6.114', '6.115', '6.116', '6.117', '6.118', '6.119',
                    '6.120', '6.122', '6.123', '6.250', '6.251', '6.252', '6.253', '6.254', '6.255', '6.256',
                    '6.257', '6.258', '6.401', '6.402', '6.403', '6.404', '6.651', '6.652', '6.653', '6.654',
                    '6.655', '6.656', '6.667', '7.100', '7.101', '7.102', '7.105', '7.106','7.127', '7.250', 
                    '7.251', '7.651', '7.654', '7.667')
    GROUP BY    
        COD_ESTABELECIMENTO,
        COD_CLIENTE,
        NOME_CLIENTE
    """
    
    # Caminho para salvar o arquivo
    caminho_base = r"C:\Users\matheus.melo\OneDrive - Acumuladores Moura SA\Documentos\Drive - Matheus Melo\Auditoria\2026\04. Circularização\Validações\Fluminense - R121"
    
    # Criar o diretório se não existir
    if not os.path.exists(caminho_base):
        os.makedirs(caminho_base)
        print(f"Diretório criado: {caminho_base}")
    
    # Conectar ao banco de dados
    conexao = conectar_banco()
    
    if conexao is None:
        print("Não foi possível conectar ao banco de dados. Verifique as credenciais.")
        return
    
    try:
        # Executar as queries
        print("Executando query de devoluções...")
        df_devolucao = executar_query(conexao, query_devolucao)
        
        print("Executando query de faturamento...")
        df_faturamento = executar_query(conexao, query_faturamento)
        
        # Verificar se os DataFrames não estão vazios
        if df_devolucao.empty or df_faturamento.empty:
            print("Atenção: Uma ou ambas as queries retornaram dados vazios.")
            if df_devolucao.empty:
                print("- Query de devolução retornou vazio.")
            if df_faturamento.empty:
                print("- Query de faturamento retornou vazio.")
        
        # Realizar o merge (join) dos dois DataFrames
        print("Consolidando dados...")
        df_consolidado = pd.merge(
            df_faturamento,
            df_devolucao,
            on=['COD_ESTABELECIMENTO', 'COD_CLIENTE', 'NOME_CLIENTE'],
            how='left'  # Left join para incluir todos os clientes com vendas
        )
        
        # Preencher valores nulos com 0 para devoluções (caso cliente não tenha devoluções)
        df_consolidado['VOLUME_DEVOLVIDO'] = df_consolidado['VOLUME_DEVOLVIDO'].fillna(0)
        df_consolidado['VALOR_DEVOLVIDO'] = df_consolidado['VALOR_DEVOLVIDO'].fillna(0)
        
        # Calcular percentual de devolução sobre vendas (com tratamento para divisão por zero)
        def calcular_percentual(valor_vendas, valor_devolvido):
            if valor_vendas == 0:
                return 0.0
            return (valor_devolvido / valor_vendas) * 100
        
        df_consolidado['%_DEVOLUCAO_VENDAS'] = df_consolidado.apply(
            lambda row: calcular_percentual(row['VALOR_VENDAS'], row['VALOR_DEVOLVIDO']),
            axis=1
        )
        
        # Arredondar valores
        df_consolidado['VALOR_VENDAS'] = df_consolidado['VALOR_VENDAS'].round(2)
        df_consolidado['VALOR_DEVOLVIDO'] = df_consolidado['VALOR_DEVOLVIDO'].round(2)
        df_consolidado['%_DEVOLUCAO_VENDAS'] = df_consolidado['%_DEVOLUCAO_VENDAS'].round(2)
        
        # Ordenar por valor de vendas (decrescente)
        df_consolidado = df_consolidado.sort_values('VALOR_VENDAS', ascending=False)
        
        # Reordenar colunas conforme solicitado
        colunas_ordenadas = [
            'COD_ESTABELECIMENTO',
            'COD_CLIENTE', 
            'NOME_CLIENTE',
            'VALOR_VENDAS',
            'VOLUME_VENDAS',
            'VALOR_DEVOLVIDO',
            'VOLUME_DEVOLVIDO',
            '%_DEVOLUCAO_VENDAS'
        ]
        df_consolidado = df_consolidado[colunas_ordenadas]
        
        # Gerar nome do arquivo
        data_atual = datetime.now().strftime('%Y%m%d')
        nome_arquivo = f'Consolidado_Vendas_Devolucoes_R121_{data_atual}.xlsx'
        caminho_completo = os.path.join(caminho_base, nome_arquivo)
        
        # Exportar para Excel
        print(f"\nExportando para: {caminho_completo}")
        
        with pd.ExcelWriter(caminho_completo, engine='openpyxl') as writer:
            df_consolidado.to_excel(writer, sheet_name='Consolidado', index=False)
            
            # Formatação da planilha
            workbook = writer.book
            worksheet = writer.sheets['Consolidado']
            
            # Formatar cabeçalhos
            header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
            
            for col in range(1, len(colunas_ordenadas) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
            
            # Formatar colunas numéricas
            number_format = '#,##0.00'
            percent_format = '0.00%'
            
            # Colunas para formato monetário
            colunas_monetarias = ['VALOR_VENDAS', 'VALOR_DEVOLVIDO']
            for col_name in colunas_monetarias:
                if col_name in df_consolidado.columns:
                    col_idx = colunas_ordenadas.index(col_name) + 1
                    for row in range(2, len(df_consolidado) + 2):
                        cell = worksheet.cell(row=row, column=col_idx)
                        cell.number_format = number_format
            
            # Formatar percentual
            if '%_DEVOLUCAO_VENDAS' in df_consolidado.columns:
                col_idx = colunas_ordenadas.index('%_DEVOLUCAO_VENDAS') + 1
                for row in range(2, len(df_consolidado) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = '0.00%'
                    # Colorir células com devolução alta (>10%)
                    if cell.value and cell.value > 10:
                        cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Ajustar largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Congelar painel (cabeçalho)
            worksheet.freeze_panes = 'A2'
        
        print("Processo concluído com sucesso!")
        
        # Estatísticas
        print(f"\n=== RESUMO DO RELATÓRIO ===")
        print(f"Total de clientes: {len(df_consolidado)}")
        print(f"Total de vendas: R$ {df_consolidado['VALOR_VENDAS'].sum():,.2f}")
        print(f"Total de devoluções: R$ {df_consolidado['VALOR_DEVOLVIDO'].sum():,.2f}")
        
        if df_consolidado['VALOR_VENDAS'].sum() > 0:
            percentual_total = (df_consolidado['VALOR_DEVOLVIDO'].sum() / df_consolidado['VALOR_VENDAS'].sum()) * 100
            print(f"Percentual total de devolução: {percentual_total:.2f}%")
        
        # Clientes com maior devolução
        print(f"\nTop 5 clientes com maior percentual de devolução:")
        top_devolucao = df_consolidado[df_consolidado['%_DEVOLUCAO_VENDAS'] > 0].nlargest(5, '%_DEVOLUCAO_VENDAS')
        for idx, row in top_devolucao.iterrows():
            print(f"  {row['NOME_CLIENTE']}: {row['%_DEVOLUCAO_VENDAS']:.2f}%")
        
        print(f"\nArquivo salvo com sucesso em:\n{caminho_completo}")
        
    except ImportError:
        print("Biblioteca openpyxl não encontrada. Instale com: pip install openpyxl")
        # Versão simplificada sem formatação
        try:
            import openpyxl
        except:
            caminho_completo = os.path.join(caminho_base, f'Consolidado_Vendas_Devolucoes_R121_{data_atual}.csv')
            df_consolidado.to_csv(caminho_completo, index=False, sep=';', decimal=',')
            print(f"Arquivo salvo em formato CSV: {caminho_completo}")
    
    except Exception as e:
        print(f"Erro durante o processamento: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        # Fechar conexão
        if conexao:
            conexao.close()
            print("\nConexão com o banco de dados fechada.")

if __name__ == "__main__":
     
    main()