import pyodbc
import pandas as pd
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# Configurações de conexão com o banco de dados
def conectar_banco():
    """
    Estabelece conexão com o banco de dados SQL Server
    Ajuste os parâmetros conforme sua configuração
    """
    try:
        # Configurações de conexão com o banco de dados
        server = '' -- Inserir o servidor  
        database = '' -- Inserir o Banco de Dados   
        username = '' -- Inserir o usuário 
        password = '' -- Inserir a senha
        
        # String de conexão
        conn_str = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'
        
        # Estabelecer conexão
        conexao = pyodbc.connect(conn_str)
        print("Conexão estabelecida com sucesso!")
        return conexao
        
    except Exception as e:
        print(f"Erro ao conectar ao banco de dados: {e}")
        return None

def executar_query(conexao, query):
    """
    Executa uma query SQL e retorna um DataFrame pandas
    """
    try:
        df = pd.read_sql(query, conexao)
        return df
    except Exception as e:
        print(f"Erro ao executar query: {e}")
        return pd.DataFrame()

def formatar_excel(caminho_arquivo, df):
    """
    Formata o arquivo Excel para melhor visualização
    """
    try:
        # Carregar o arquivo Excel
        workbook = openpyxl.load_workbook(caminho_arquivo)
        worksheet = workbook.active
        
        # Definir estilos
        header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Vermelho escuro
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Formatar cabeçalhos
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Formatar colunas monetárias
        colunas_monetarias = ['VALOR_VENDAS', 'VALOR_DEVOLVIDO_TERCEIROS']
        number_format = '#,##0.00'
        
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in colunas_monetarias:
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = number_format
        
        # Formatar percentual
        if '%_DEVOLUCAO_TERCEIROS' in df.columns:
            col_idx = list(df.columns).index('%_DEVOLUCAO_TERCEIROS') + 1
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = '0.00%'
                # Destacar devoluções altas (>10%)
                if cell.value and cell.value > 0.10:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                # Destacar devoluções muito altas (>20%)
                if cell.value and cell.value > 0.20:
                    cell.font = Font(color="FF0000", bold=True)
        
        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar painel (cabeçalho)
        worksheet.freeze_panes = 'A2'
        
        # Adicionar filtros automáticos
        worksheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}1"
        
        # Salvar as alterações
        workbook.save(caminho_arquivo)
        print(f"Formatação aplicada ao arquivo: {caminho_arquivo}")
        
    except Exception as e:
        print(f"Erro ao formatar Excel: {e}")

def gerar_relatorio_devolucoes_nf_propria():
    """
    Gera relatório de devoluções por NF própria (terceiros)
    """
    # Definir as queries - ATENÇÃO: Períodos diferentes nas queries!
    query_devolucao_nf_propria = """
    SELECT
        COD_ESTABELECIMENTO,
        COD_CLIENTE,
        NOME_CLIENTE,
        SUM(QUANTIDADE) AS VOLUME_DEVOLVIDO_TERCEIROS,
        SUM(VALOR) AS VALOR_DEVOLVIDO_TERCEIROS
    FROM 
        VW_AUDIT_RM_ORDENS_VENDA
    WHERE
        COD_ESTABELECIMENTO = 'R121'
        AND DATA_NOTA_FISCAL BETWEEN '2025-07-01' AND '2025-12-31' 
        AND PARA_FATURAMENTO = 'SIM'
        AND EMISSOR = 'OwnEstablishment'
        AND CFOP IN ('1.201', '1.202', '1.203', '1.204', '1.410', '1.411', '1.553', '1.660', '1.661', '1.662', 
                    '2.201', '2.202', '2.203', '2.204', '2.410', '2.411', '2.553', '2.660', '2.661', '2.662',
                    '3.201', '3.202', '3.211', '3.553')
    GROUP BY
        COD_ESTABELECIMENTO,
        COD_CLIENTE,
        NOME_CLIENTE
    """

    query_faturamento = """
    SELECT
        COD_ESTABELECIMENTO,
        COD_CLIENTE,
        NOME_CLIENTE,
        SUM(QUANTIDADE) AS VOLUME_VENDAS,
        SUM(VALOR) AS VALOR_VENDAS
    FROM 
        VW_AUDIT_RM_ORDENS_VENDA
    WHERE 
        COD_ESTABELECIMENTO = 'R121' 
        AND DATA_NOTA_FISCAL BETWEEN '2025-07-01' AND '2025-12-31'  
        AND PARA_FATURAMENTO = 'Sim'
        AND CFOP IN ('5.100', '5.101', '5.102', '5.103', '5.104', '5.105', '5.106', '5.109', '5.110', '5.111', 
                    '5.112', '5.113', '5.114', '5.115', '5.116', '5.117', '5.118', '5.119', '5.120', '5.122', 
                    '5.123', '5.250','5.251', '5.252', '5.253', '5.254', '5.255', '5.256', '5.257', '5.258', 
                    '5.401', '5.402', '5.403', '5.405', '5.651', '5.652', '5.653', '5.654', '5.655', '5.656',
                    '5.667', '6.101', '6.102', '6.103','6.104', '6.105', '6.106', '6.107', '6.108', '6.109',
                    '6.110', '6.111', '6.112', '6.113', '6.114', '6.115', '6.116', '6.117', '6.118', '6.119',
                    '6.120', '6.122', '6.123', '6.250', '6.251', '6.252', '6.253', '6.254', '6.255', '6.256',
                    '6.257', '6.258', '6.401', '6.402', '6.403', '6.404', '6.651', '6.652', '6.653', '6.654',
                    '6.655', '6.656', '6.667', '7.100', '7.101', '7.102', '7.105', '7.106','7.127', '7.250', 
                    '7.251', '7.651', '7.654', '7.667')
    GROUP BY    
        COD_ESTABELECIMENTO,
        COD_CLIENTE,
        NOME_CLIENTE
    """
    
    # Caminho para salvar o arquivo
    caminho_base = r"C:\Users\matheus.melo\OneDrive - Acumuladores Moura SA\Documentos\Drive - Matheus Melo\Auditoria\2026\04. Circularização\Validações\Fluminense - R121"
    
    # Criar o diretório se não existir
    if not os.path.exists(caminho_base):
        os.makedirs(caminho_base)
        print(f"Diretório criado: {caminho_base}")
    
    # Conectar ao banco de dados
    conexao = conectar_banco()
    
    if conexao is None:
        print("Não foi possível conectar ao banco de dados. Verifique as credenciais.")
        return None, None
    
    try:
        # Executar as queries
        print("Executando query de devoluções por NF própria...")
        df_devolucao = executar_query(conexao, query_devolucao_nf_propria)
        
        print("Executando query de faturamento...")
        df_faturamento = executar_query(conexao, query_faturamento)
        
        # Verificar se os DataFrames não estão vazios
        if df_faturamento.empty:
            print("ERRO: Query de faturamento retornou vazio.")
            return None, None
        
        if df_devolucao.empty:
            print("AVISO: Query de devoluções por NF própria retornou vazio (nenhuma devolução encontrada).")
            # Criar DataFrame vazio para devoluções
            df_devolucao = pd.DataFrame(columns=['COD_ESTABELECIMENTO', 'COD_CLIENTE', 'NOME_CLIENTE', 
                                                 'VOLUME_DEVOLVIDO_TERCEIROS', 'VALOR_DEVOLVIDO_TERCEIROS'])
        
        # AVISO: Períodos diferentes nas queries
        print("\n⚠️  ATENÇÃO: As queries usam períodos diferentes!")
        print(f"   • Devoluções NF própria: 01/07/2025 a 31/12/2025")
        print(f"   • Faturamento: 07/07/2025 a 07/01/2026")
        print("   (Esta diferença pode afetar a análise comparativa)")
        
        # Realizar o merge (join) dos dois DataFrames
        print("\nConsolidando dados...")
        df_consolidado = pd.merge(
            df_faturamento,
            df_devolucao,
            on=['COD_ESTABELECIMENTO', 'COD_CLIENTE', 'NOME_CLIENTE'],
            how='left'  # Left join para incluir todos os clientes com vendas
        )
        
        # Preencher valores nulos com 0 para devoluções (caso cliente não tenha devoluções)
        df_consolidado['VOLUME_DEVOLVIDO_TERCEIROS'] = df_consolidado['VOLUME_DEVOLVIDO_TERCEIROS'].fillna(0)
        df_consolidado['VALOR_DEVOLVIDO_TERCEIROS'] = df_consolidado['VALOR_DEVOLVIDO_TERCEIROS'].fillna(0)
        
        # Calcular percentual de devoluções sobre vendas (com tratamento para divisão por zero)
        def calcular_percentual(valor_vendas, valor_devolvido):
            if valor_vendas == 0:
                return 0.0
            return (valor_devolvido / valor_vendas)
        
        df_consolidado['%_DEVOLUCAO_TERCEIROS'] = df_consolidado.apply(
            lambda row: calcular_percentual(row['VALOR_VENDAS'], row['VALOR_DEVOLVIDO_TERCEIROS']),
            axis=1
        )
        
        # Arredondar valores
        df_consolidado['VALOR_VENDAS'] = df_consolidado['VALOR_VENDAS'].round(2)
        df_consolidado['VALOR_DEVOLVIDO_TERCEIROS'] = df_consolidado['VALOR_DEVOLVIDO_TERCEIROS'].round(2)
        df_consolidado['VOLUME_VENDAS'] = df_consolidado['VOLUME_VENDAS'].round(0).astype(int)
        df_consolidado['VOLUME_DEVOLVIDO_TERCEIROS'] = df_consolidado['VOLUME_DEVOLVIDO_TERCEIROS'].round(0).astype(int)
        
        # Ordenar por valor de vendas (decrescente)
        df_consolidado = df_consolidado.sort_values('VALOR_VENDAS', ascending=False)
        
        # Reordenar colunas conforme solicitado
        colunas_ordenadas = [
            'COD_ESTABELECIMENTO',
            'COD_CLIENTE', 
            'NOME_CLIENTE',
            'VALOR_VENDAS',
            'VOLUME_VENDAS',
            'VALOR_DEVOLVIDO_TERCEIROS',
            'VOLUME_DEVOLVIDO_TERCEIROS',
            '%_DEVOLUCAO_TERCEIROS'
        ]
        
        # Garantir que todas as colunas existem
        df_consolidado = df_consolidado.reindex(columns=colunas_ordenadas)
        
        # Gerar nome do arquivo
        data_atual = datetime.now().strftime('%Y%m%d_%H%M')
        nome_arquivo = f'Devolucoes_NF_Propria_R121_{data_atual}.xlsx'
        caminho_completo = os.path.join(caminho_base, nome_arquivo)
        
        # Exportar para Excel
        print(f"\nExportando para: {caminho_completo}")
        df_consolidado.to_excel(caminho_completo, sheet_name='Devolucoes_NF_Propria', index=False)
        
        # Formatar o Excel
        formatar_excel(caminho_completo, df_consolidado)
        
        print("\n" + "="*70)
        print("RELATÓRIO DE DEVOLUÇÕES POR NF PRÓPRIA - CONCLUÍDO!")
        print("="*70)
        
        # Estatísticas detalhadas
        print(f"\n=== RESUMO ESTATÍSTICO ===")
        print(f"Período de análise:")
        print(f"   • Devoluções (NF própria): 01/07/2025 a 31/12/2025")
        print(f"   • Faturamento: 07/07/2025 a 07/01/2026")
        print(f"Estabelecimento: R121")
        print(f"\nTotal de clientes no período: {len(df_consolidado)}")
        print(f"Clientes com devoluções por NF própria: {(df_consolidado['VALOR_DEVOLVIDO_TERCEIROS'] > 0).sum()}")
        
        # Valores totais
        total_vendas = df_consolidado['VALOR_VENDAS'].sum()
        total_devolucoes = df_consolidado['VALOR_DEVOLVIDO_TERCEIROS'].sum()
        
        print(f"\nValor total de vendas: R$ {total_vendas:,.2f}")
        print(f"Valor total devolvido (NF própria): R$ {total_devolucoes:,.2f}")
        
        if total_vendas > 0:
            percentual_total = (total_devolucoes / total_vendas) * 100
            print(f"Percentual total de devoluções (NF própria): {percentual_total:.2f}%")
        
        # Análise dos clientes com devoluções
        clientes_com_devolucao = df_consolidado[df_consolidado['VALOR_DEVOLVIDO_TERCEIROS'] > 0]
        
        if not clientes_com_devolucao.empty:
            print(f"\n=== CLIENTES COM DEVOLUÇÕES POR NF PRÓPRIA ===")
            print(f"Total: {len(clientes_com_devolucao)} clientes")
            
            # Top 5 clientes com maior valor devolvido
            print(f"\nTop 5 clientes - Maior valor devolvido (NF própria):")
            top_valor = clientes_com_devolucao.nlargest(5, 'VALOR_DEVOLVIDO_TERCEIROS')
            for idx, row in top_valor.iterrows():
                percentual = (row['VALOR_DEVOLVIDO_TERCEIROS'] / row['VALOR_VENDAS']) * 100 if row['VALOR_VENDAS'] > 0 else 0
                print(f"  {row['NOME_CLIENTE']}: R$ {row['VALOR_DEVOLVIDO_TERCEIROS']:,.2f} ({percentual:.1f}% das vendas)")
            
            # Top 5 clientes com maior percentual de devolução
            print(f"\nTop 5 clientes - Maior percentual de devolução (NF própria):")
            clientes_com_vendas = clientes_com_devolucao[clientes_com_devolucao['VALOR_VENDAS'] > 0]
            if len(clientes_com_vendas) > 0:
                top_percentual = clientes_com_vendas.nlargest(5, '%_DEVOLUCAO_TERCEIROS')
                for idx, row in top_percentual.iterrows():
                    percentual = row['%_DEVOLUCAO_TERCEIROS'] * 100
                    print(f"  {row['NOME_CLIENTE']}: {percentual:.1f}% (R$ {row['VALOR_DEVOLVIDO_TERCEIROS']:,.2f})")
            
            # Estatísticas de distribuição
            print(f"\nDistribuição de devoluções por faixa percentual:")
            faixas = [
                (0, 0.05, "Até 5%"),
                (0.05, 0.10, "5% a 10%"),
                (0.10, 0.20, "10% a 20%"),
                (0.20, 1.0, "Acima de 20%")
            ]
            
            for min_val, max_val, label in faixas:
                count = len(clientes_com_vendas[
                    (clientes_com_vendas['%_DEVOLUCAO_TERCEIROS'] >= min_val) & 
                    (clientes_com_vendas['%_DEVOLUCAO_TERCEIROS'] < max_val)
                ])
                if count > 0:
                    print(f"  {label}: {count} cliente(s)")
        
        # Análise dos 10 principais clientes por faturamento
        print(f"\n=== TOP 10 CLIENTES POR FATURAMENTO ===")
        top_10_vendas = df_consolidado.head(10)
        for idx, row in top_10_vendas.iterrows():
            tem_devolucao = "SIM" if row['VALOR_DEVOLVIDO_TERCEIROS'] > 0 else "NÃO"
            percentual = (row['VALOR_DEVOLVIDO_TERCEIROS'] / row['VALOR_VENDAS']) * 100 if row['VALOR_VENDAS'] > 0 else 0
            print(f"  {row['NOME_CLIENTE']}: R$ {row['VALOR_VENDAS']:,.2f} | Devolução NF própria: {tem_devolucao} | %: {percentual:.1f}%")
        
        print(f"\n📍 Arquivo salvo em:")
        print(f"   {caminho_completo}")
        print(f"\n📊 Campos incluídos no relatório:")
        for coluna in colunas_ordenadas:
            print(f"   • {coluna}")
        
        return caminho_completo, df_consolidado
        
    except Exception as e:
        print(f"\n❌ ERRO durante o processamento: {e}")
        import traceback
        traceback.print_exc()
        return None, None
    
    finally:
        # Fechar conexão
        if conexao:
            conexao.close()
            print("\nConexão com o banco de dados fechada.")

def main():
    """
    Função principal para executar o relatório
    """
    print("="*70)
    print("RELATÓRIO DE DEVOLUÇÕES POR NOTA FISCAL PRÓPRIA (TERCEIROS) - R121")
    print("="*70)
    print("\nIniciando processo de geração do relatório...")
    
    # Verificar dependências
    try:
        import pyodbc
        import pandas
        import openpyxl
        print("✅ Todas as dependências estão instaladas")
    except ImportError as e:
        print(f"\n❌ Falta instalar dependências: {e}")
        print("\nInstale com:")
        print("   pip install pyodbc pandas openpyxl")
        return
    
    # Executar o relatório
    caminho_arquivo, df_resultado = gerar_relatorio_devolucoes_nf_propria()
    
    if caminho_arquivo:
        print("\n" + "="*70)
        print("✅ RELATÓRIO GERADO COM SUCESSO!")
        print("="*70)
        
        # Oferecer opção para visualizar primeiras linhas
        resposta = input("\nDeseja visualizar as primeiras 10 linhas do relatório? (S/N): ")
        if resposta.upper() == 'S' and df_resultado is not None:
            print("\nPrimeiras 10 linhas do relatório:")
            print(df_resultado.head(10).to_string(index=False))

if __name__ == "__main__":
    main()

