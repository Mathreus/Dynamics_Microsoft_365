import pyodbc
import pandas as pd
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# Configurações de conexão com o banco de dados
def conectar_banco():
    """
    Estabelece conexão com o banco de dados SQL Server
    Ajuste os parâmetros conforme sua configuração
    """
    try:
        # Configurações de conexão - AJUSTE AQUI CONFORME SEU AMBIENTE
        server = 'DCMDWF01A.MOURA.INT'  # Ex: 'localhost\\SQLEXPRESS' ou 'dcmdwf01a.moura.int'
        database = 'ax'   # Ex: 'SeuDatabase'
        username = 'uAuditoria' # Ex: 'sa'
        password = '@ud!t0$!@202&22'   # Ex: 'sua_senha'
        
        # String de conexão
        conn_str = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'
        
        # Estabelecer conexão
        conexao = pyodbc.connect(conn_str)
        print("Conexão estabelecida com sucesso!")
        return conexao
        
    except Exception as e:
        print(f"Erro ao conectar ao banco de dados: {e}")
        return None

def executar_query(conexao, query):
    """
    Executa uma query SQL e retorna um DataFrame pandas
    """
    try:
        df = pd.read_sql(query, conexao)
        return df
    except Exception as e:
        print(f"Erro ao executar query: {e}")
        return pd.DataFrame()

def formatar_excel(caminho_arquivo, df):
    """
    Formata o arquivo Excel para melhor visualização
    """
    try:
        # Carregar o arquivo Excel
        workbook = openpyxl.load_workbook(caminho_arquivo)
        worksheet = workbook.active
        
        # Definir estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Formatar cabeçalhos
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Formatar colunas monetárias
        colunas_monetarias = ['VALOR_VENDAS', 'VALOR_BONIFICADO']
        number_format = '#,##0.00'
        
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in colunas_monetarias:
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = number_format
        
        # Formatar percentual
        if '%_BONIFICACOES' in df.columns:
            col_idx = list(df.columns).index('%_BONIFICACOES') + 1
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = '0.00%'
                # Destacar bonificações altas (>15%)
                if cell.value and cell.value > 0.15:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar painel (cabeçalho)
        worksheet.freeze_panes = 'A2'
        
        # Salvar as alterações
        workbook.save(caminho_arquivo)
        print(f"Formatação aplicada ao arquivo: {caminho_arquivo}")
        
    except Exception as e:
        print(f"Erro ao formatar Excel: {e}")

def gerar_relatorio_bonificacoes():
    """
    Gera relatório de bonificações por cliente
    """
    # Definir as queries
    query_bonificacao = """
    SELECT
        COD_ESTABELECIMENTO,
        COD_CLIENTE,
        NOME_CLIENTE,
        SUM(QUANTIDADE) AS VOLUME_BONIFICADO,
        SUM(VALOR) AS VALOR_BONIFICADO
    FROM 
        VW_AUDIT_RM_ORDENS_VENDA
    WHERE
        COD_ESTABELECIMENTO = 'R121'
        AND DATA_NOTA_FISCAL BETWEEN '2025-07-07' AND '2026-01-07' 
        AND DESC_TIPO_OPERACAO LIKE '%REMESSA EM BONIFICACAO%'
    GROUP BY
        COD_ESTABELECIMENTO,
        COD_CLIENTE,
        NOME_CLIENTE
    """

    query_faturamento = """
    SELECT
        COD_ESTABELECIMENTO,
        COD_CLIENTE,
        NOME_CLIENTE,
        SUM(QUANTIDADE) AS VOLUME_VENDAS,
        SUM(VALOR) AS VALOR_VENDAS
    FROM 
        VW_AUDIT_RM_ORDENS_VENDA
    WHERE 
        COD_ESTABELECIMENTO = 'R121' 
        AND DATA_NOTA_FISCAL BETWEEN '2025-07-07' AND '2026-01-07'  
        AND PARA_FATURAMENTO = 'Sim'
        AND CFOP IN ('5.100', '5.101', '5.102', '5.103', '5.104', '5.105', '5.106', '5.109', '5.110', '5.111', 
                    '5.112', '5.113', '5.114', '5.115', '5.116', '5.117', '5.118', '5.119', '5.120', '5.122', 
                    '5.123', '5.250','5.251', '5.252', '5.253', '5.254', '5.255', '5.256', '5.257', '5.258', 
                    '5.401', '5.402', '5.403', '5.405', '5.651', '5.652', '5.653', '5.654', '5.655', '5.656',
                    '5.667', '6.101', '6.102', '6.103','6.104', '6.105', '6.106', '6.107', '6.108', '6.109',
                    '6.110', '6.111', '6.112', '6.113', '6.114', '6.115', '6.116', '6.117', '6.118', '6.119',
                    '6.120', '6.122', '6.123', '6.250', '6.251', '6.252', '6.253', '6.254', '6.255', '6.256',
                    '6.257', '6.258', '6.401', '6.402', '6.403', '6.404', '6.651', '6.652', '6.653', '6.654',
                    '6.655', '6.656', '6.667', '7.100', '7.101', '7.102', '7.105', '7.106','7.127', '7.250', 
                    '7.251', '7.651', '7.654', '7.667')
    GROUP BY    
        COD_ESTABELECIMENTO,
        COD_CLIENTE,
        NOME_CLIENTE
    """
    
    # Caminho para salvar o arquivo
    caminho_base = r"C:\Users\matheus.melo\OneDrive - Acumuladores Moura SA\Documentos\Drive - Matheus Melo\Auditoria\2026\04. Circularização\Validações\Fluminense - R121"
    
    # Criar o diretório se não existir
    if not os.path.exists(caminho_base):
        os.makedirs(caminho_base)
        print(f"Diretório criado: {caminho_base}")
    
    # Conectar ao banco de dados
    conexao = conectar_banco()
    
    if conexao is None:
        print("Não foi possível conectar ao banco de dados. Verifique as credenciais.")
        return
    
    try:
        # Executar as queries
        print("Executando query de bonificações...")
        df_bonificacao = executar_query(conexao, query_bonificacao)
        
        print("Executando query de faturamento...")
        df_faturamento = executar_query(conexao, query_faturamento)
        
        # Verificar se os DataFrames não estão vazios
        if df_faturamento.empty:
            print("ERRO: Query de faturamento retornou vazio.")
            return
        
        if df_bonificacao.empty:
            print("AVISO: Query de bonificações retornou vazio (nenhuma bonificação encontrada).")
            # Criar DataFrame vazio para bonificações
            df_bonificacao = pd.DataFrame(columns=['COD_ESTABELECIMENTO', 'COD_CLIENTE', 'NOME_CLIENTE', 
                                                   'VOLUME_BONIFICADO', 'VALOR_BONIFICADO'])
        
        # Realizar o merge (join) dos dois DataFrames
        print("Consolidando dados...")
        df_consolidado = pd.merge(
            df_faturamento,
            df_bonificacao,
            on=['COD_ESTABELECIMENTO', 'COD_CLIENTE', 'NOME_CLIENTE'],
            how='left'  # Left join para incluir todos os clientes com vendas
        )
        
        # Preencher valores nulos com 0 para bonificações (caso cliente não tenha bonificações)
        df_consolidado['VOLUME_BONIFICADO'] = df_consolidado['VOLUME_BONIFICADO'].fillna(0)
        df_consolidado['VALOR_BONIFICADO'] = df_consolidado['VALOR_BONIFICADO'].fillna(0)
        
        # Calcular percentual de bonificações sobre vendas (com tratamento para divisão por zero)
        def calcular_percentual(valor_vendas, valor_bonificado):
            if valor_vendas == 0:
                return 0.0
            return (valor_bonificado / valor_vendas)
        
        df_consolidado['%_BONIFICACOES'] = df_consolidado.apply(
            lambda row: calcular_percentual(row['VALOR_VENDAS'], row['VALOR_BONIFICADO']),
            axis=1
        )
        
        # Arredondar valores
        df_consolidado['VALOR_VENDAS'] = df_consolidado['VALOR_VENDAS'].round(2)
        df_consolidado['VALOR_BONIFICADO'] = df_consolidado['VALOR_BONIFICADO'].round(2)
        df_consolidado['VOLUME_VENDAS'] = df_consolidado['VOLUME_VENDAS'].round(0).astype(int)
        df_consolidado['VOLUME_BONIFICADO'] = df_consolidado['VOLUME_BONIFICADO'].round(0).astype(int)
        
        # Ordenar por valor de vendas (decrescente)
        df_consolidado = df_consolidado.sort_values('VALOR_VENDAS', ascending=False)
        
        # Reordenar colunas conforme solicitado
        colunas_ordenadas = [
            'COD_ESTABELECIMENTO',
            'COD_CLIENTE', 
            'NOME_CLIENTE',
            'VALOR_VENDAS',
            'VOLUME_VENDAS',
            'VALOR_BONIFICADO',
            'VOLUME_BONIFICADO',
            '%_BONIFICACOES'
        ]
        
        # Garantir que todas as colunas existem
        df_consolidado = df_consolidado.reindex(columns=colunas_ordenadas)
        
        # Gerar nome do arquivo
        data_atual = datetime.now().strftime('%Y%m%d_%H%M')
        nome_arquivo = f'Bonificacoes_Vendas_R121_{data_atual}.xlsx'
        caminho_completo = os.path.join(caminho_base, nome_arquivo)
        
        # Exportar para Excel
        print(f"\nExportando para: {caminho_completo}")
        df_consolidado.to_excel(caminho_completo, sheet_name='Bonificacoes', index=False)
        
        # Formatar o Excel
        formatar_excel(caminho_completo, df_consolidado)
        
        print("\n" + "="*60)
        print("PROCESSO CONCLUÍDO COM SUCESSO!")
        print("="*60)
        
        # Estatísticas
        print(f"\n=== RESUMO DO RELATÓRIO DE BONIFICAÇÕES ===")
        print(f"Período: 07/07/2025 a 07/01/2026")
        print(f"Estabelecimento: R121")
        print(f"Total de clientes: {len(df_consolidado)}")
        print(f"Clientes com bonificações: {(df_consolidado['VALOR_BONIFICADO'] > 0).sum()}")
        print(f"\nValor total de vendas: R$ {df_consolidado['VALOR_VENDAS'].sum():,.2f}")
        print(f"Valor total bonificado: R$ {df_consolidado['VALOR_BONIFICADO'].sum():,.2f}")
        
        if df_consolidado['VALOR_VENDAS'].sum() > 0:
            percentual_total = (df_consolidado['VALOR_BONIFICADO'].sum() / df_consolidado['VALOR_VENDAS'].sum()) * 100
            print(f"Percentual total de bonificações: {percentual_total:.2f}%")
        
        # Top 5 clientes com maior bonificação
        clientes_com_bonificacao = df_consolidado[df_consolidado['VALOR_BONIFICADO'] > 0]
        if not clientes_com_bonificacao.empty:
            print(f"\nTop 5 clientes com maior valor bonificado:")
            top_bonificacao = clientes_com_bonificacao.nlargest(5, 'VALOR_BONIFICADO')
            for idx, row in top_bonificacao.iterrows():
                percentual = (row['VALOR_BONIFICADO'] / row['VALOR_VENDAS']) * 100 if row['VALOR_VENDAS'] > 0 else 0
                print(f"  {row['NOME_CLIENTE']}: R$ {row['VALOR_BONIFICADO']:,.2f} ({percentual:.1f}%)")
        
        # Top 5 clientes com maior percentual de bonificação
        clientes_com_vendas = df_consolidado[df_consolidado['VALOR_VENDAS'] > 0]
        if len(clientes_com_vendas) > 0:
            print(f"\nTop 5 clientes com maior percentual de bonificação:")
            top_percentual = clientes_com_vendas[clientes_com_vendas['%_BONIFICACOES'] > 0].nlargest(5, '%_BONIFICACOES')
            for idx, row in top_percentual.iterrows():
                percentual = row['%_BONIFICACOES'] * 100
                print(f"  {row['NOME_CLIENTE']}: {percentual:.1f}% (R$ {row['VALOR_BONIFICADO']:,.2f} / R$ {row['VALOR_VENDAS']:,.2f})")
        
        print(f"\nArquivo salvo com sucesso em:")
        print(f"{caminho_completo}")
        print("\nCampos incluídos:")
        for coluna in colunas_ordenadas:
            print(f"  - {coluna}")
        
        return caminho_completo, df_consolidado
        
    except Exception as e:
        print(f"\nERRO durante o processamento: {e}")
        import traceback
        traceback.print_exc()
        return None, None
    
    finally:
        # Fechar conexão
        if conexao:
            conexao.close()
            print("\nConexão com o banco de dados fechada.")

def main():
    """
    Função principal para executar o relatório
    """
    print("="*60)
    print("RELATÓRIO DE BONIFICAÇÕES POR CLIENTE - R121")
    print("="*60)
    print("\nIniciando processo de geração do relatório...")
    
    # Verificar dependências
    try:
        import pyodbc
        import pandas
        import openpyxl
        print("✓ Todas as dependências estão instaladas")
    except ImportError as e:
        print(f"\n✗ Falta instalar dependências: {e}")
        print("\nInstale com:")
        print("pip install pyodbc pandas openpyxl")
        return
    
    # Executar o relatório
    caminho_arquivo, df_resultado = gerar_relatorio_bonificacoes()
    
    if caminho_arquivo:
        print("\n" + "="*60)
        print("PRONTO! O relatório foi gerado com sucesso.")
        print("="*60)

if __name__ == "__main__":
    main()
